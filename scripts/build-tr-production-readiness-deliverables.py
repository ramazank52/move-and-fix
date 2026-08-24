from pathlib import Path
import csv
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
EVIDENCE = ROOT / "docs" / "compliance" / "tr-production-evidence"
EXPORTS = ROOT / "exports"
EXPORTS.mkdir(exist_ok=True)

def workbook_from_csv(csv_name: str, xlsx_name: str, sheet_name: str):
    with (EVIDENCE / csv_name).open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in col) + 2, 12), 55)
    wb.save(EXPORTS / xlsx_name)

workbook_from_csv("TR_SERVICE_CREDENTIAL_MATRIX.csv", "TR_SERVICE_CREDENTIAL_MATRIX.xlsx", "Matrix")
workbook_from_csv("TR_LEGAL_SOURCE_COUNSEL_REVIEW_TEMPLATE.csv", "TR_LEGAL_SOURCE_COUNSEL_REVIEW_TEMPLATE.xlsx", "Counsel Review")
workbook_from_csv("TR_OFFICIAL_SOURCE_REGISTER.csv", "TR_CONNECTOR_AND_EXTERNAL_INPUT_REGISTER.xlsx", "Official Sources")
wb_path = EXPORTS / "TR_CONNECTOR_AND_EXTERNAL_INPUT_REGISTER.xlsx"
from openpyxl import load_workbook
wb = load_workbook(wb_path)
ws = wb.create_sheet("External Inputs")
ws.append(["dependency", "production input", "current state", "blocking state", "required user action", "secret handling"])
for row in [
    ["Payment provider", "IYZICO_API_KEY / IYZICO_SECRET_KEY or approved provider credentials", "NOT_CONFIGURED", "BLOCKED_EXTERNAL_INPUT", "Provide production account credentials and webhook acceptance evidence", "Set only through secure secret management"],
    ["SMS", "NETGSM or approved SMS credentials", "NOT_CONFIGURED", "BLOCKED_EXTERNAL_INPUT", "Provide sender/account credentials and delivery acceptance", "Do not place in source or reports"],
    ["Email", "Approved transactional email credentials", "NOT_CONFIGURED", "BLOCKED_EXTERNAL_INPUT", "Provide domain/provider credentials and SPF/DKIM acceptance", "Do not place in source or reports"],
    ["Push", "Apple/Google push credentials", "NOT_CONFIGURED", "BLOCKED_EXTERNAL_INPUT", "Provide APNs/FCM production configuration", "Do not place in source or reports"],
    ["Malware scanner", "Production scanner adapter credential/endpoint", "NOT_CONFIGURED", "BLOCKED_EXTERNAL_INPUT", "Provide approved scanner integration and callback test", "Do not place in source or reports"],
    ["Official document connectors", "Issuer/API contract or registered manual verification authority", "PENDING", "BLOCKED_CONNECTOR", "Provide authorization, contract, permitted method and test evidence", "No portal scraping/e-Devlet impersonation"],
    ["KMS/secret manager", "Production key-management/secret-store configuration", "NOT_CONFIGURED", "BLOCKED_EXTERNAL_INPUT", "Provide approved production KMS and rotation evidence", "No generated or default secret"],
    ["DNS/HTTPS", "Production domain, DNS and certificate control", "NOT_CONFIGURED", "BLOCKED_EXTERNAL_INPUT", "Configure domain/certificate and evidence", "No publish performed"],
    ["Apple/Google stores", "Store account, signing and release permissions", "NOT_CONFIGURED", "BLOCKED_PHYSICAL_E2E", "Provide accounts/signing only at approved release gate", "No store release performed"],
]: ws.append(row)
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = min(max(max(len(str(c.value or "")) for c in col) + 2, 12), 55)
wb.save(wb_path)
workbook_from_csv("TR_PRODUCTION_ALLOWLIST_AND_NOGO.csv", "TR_PRODUCTION_ALLOWLIST_AND_NOGO.xlsx", "Capability Decisions")
