from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
AUDIT = ROOT / "exports" / "us-ca-la-checkpoint-a-audit"
EXPORTS = ROOT / "exports"


def write_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, header in enumerate(headers, start=1):
        max_length = max([len(str(header))] + [len(str(row[column_index - 1] if column_index - 1 < len(row) else "")) for row in rows])
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 14), 48)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    matrix_path = AUDIT / "US_CA_LA_SERVICE_CREDENTIAL_MATRIX.csv"
    sources_path = AUDIT / "US_CA_LA_SOURCE_REGISTER.csv"
    bundle_path = AUDIT / "US_CA_LA_REQUIREMENT_BUNDLE_MANIFEST.json"
    connector_path = AUDIT / "US_CA_LA_CONNECTOR_INVENTORY.json"
    locale_path = AUDIT / "US_CA_LA_LOCALE_REGISTER.json"

    with matrix_path.open(newline="", encoding="utf-8") as matrix_file:
        matrix_reader = csv.DictReader(matrix_file)
        matrix_rows = list(matrix_reader)
        matrix_headers = matrix_reader.fieldnames or []
    with sources_path.open(newline="", encoding="utf-8") as sources_file:
        source_reader = csv.DictReader(sources_file)
        source_rows = list(source_reader)
        source_headers = source_reader.fieldnames or []
    bundles = json.loads(bundle_path.read_text(encoding="utf-8"))
    connectors = json.loads(connector_path.read_text(encoding="utf-8"))
    locales = json.loads(locale_path.read_text(encoding="utf-8"))

    matrix_book = Workbook()
    matrix_book.remove(matrix_book.active)
    write_sheet(matrix_book, "62 Coverage Matrix", matrix_headers, [[row.get(header, "") for header in matrix_headers] for row in matrix_rows])
    bundle_headers = ["bundleKey", "title", "riskLevel", "triggerDescription", "verificationDescription", "requiredEvidenceJson", "subjectBindings", "sourceBindings", "boundCoverageRows", "decisionIfMissing", "sourceState", "legalState", "orphan", "ambiguousAlias"]
    bundle_rows = [[
        bundle.get("bundleKey", ""), bundle.get("title", ""), bundle.get("riskLevel", ""), bundle.get("triggerDescription", ""),
        bundle.get("verificationDescription", ""), json.dumps(bundle.get("requiredEvidenceJson", []), ensure_ascii=False),
        json.dumps(bundle.get("subjectBindings", []), ensure_ascii=False), json.dumps(bundle.get("sourceBindings", []), ensure_ascii=False),
        json.dumps(bundle.get("boundCoverageRows", []), ensure_ascii=False), bundle.get("decisionIfMissing", ""), bundle.get("sourceState", ""),
        bundle.get("legalState", ""), bundle.get("orphan", ""), bundle.get("ambiguousAlias", ""),
    ] for bundle in bundles]
    write_sheet(matrix_book, "26 Requirement Bundles", bundle_headers, bundle_rows)
    connector_headers = ["connectorKey", "displayName", "officialSourceKey", "targetRegistryOrApi", "status", "assuranceLevel", "forbiddenScraping", "accessPermissionRequirement", "returnedFields", "expirySuspensionRevocationCheck", "noGoReason"]
    write_sheet(matrix_book, "28 Connector Inventory", connector_headers, [[connector.get(header, "") for header in connector_headers] for connector in connectors])
    locale_headers = ["document_key", "document_version", "document_surface", "legal_approval_state", "locale", "localization_state", "content_hash", "content_storage_key", "runtime_selectable"]
    write_sheet(matrix_book, "12 Draft Machine Locales", locale_headers, [[
        locale.get("documentKey", ""), locale.get("documentVersion", ""), locale.get("documentSurface", ""), locale.get("legalApprovalState", ""),
        locale.get("locale", ""), locale.get("localizationState", ""), locale.get("contentHash", ""), locale.get("contentStorageKey", ""), locale.get("runtimeSelectable", ""),
    ] for locale in locales])
    matrix_book.save(EXPORTS / "US_CA_LA_SERVICE_CREDENTIAL_MATRIX.xlsx")

    source_book = Workbook()
    source_book.remove(source_book.active)
    write_sheet(source_book, "28 Source Register", source_headers, [[row.get(header, "") for header in source_headers] for row in source_rows])
    source_book.save(EXPORTS / "US_CA_LA_SOURCE_REGISTER.xlsx")

    review_book = Workbook()
    review_book.remove(review_book.active)
    write_sheet(
        review_book,
        "Review Instructions",
        ["field", "value"],
        [
            ["status", "CHECKPOINT_A_REVIEW_REQUIRED"],
            ["rule", "No SOURCE_VERIFIED, LEGAL_APPROVED, connector authorization or product release decision may be entered without independent authorized evidence."],
            ["required_counsel_evidence", "Official title, section/classification/exception, effective date, jurisdiction applicability, decision, reviewer identity/role, dated evidence hash and immutable approval reference."],
            ["runtime", "All US capability coverage remains BLOCKED; this workbook is not an approval instrument by itself."],
        ],
    )
    coverage_review_headers = [
        "research_row_id", "canonical_family", "canonical_subservice", "exact_task_profile", "risk_level", "mandatory_bundles", "conditional_bundles", "subject_bindings", "current_policy_decision", "current_source_state", "current_legal_state", "counsel_decision_blank", "counsel_effective_date_blank", "counsel_exception_blank", "counsel_reviewer_identity_blank", "counsel_role_scope_blank", "counsel_evidence_hash_blank", "immutable_approval_ledger_ref_blank",
    ]
    coverage_review_rows = [[
        row.get("research_row_id", ""), row.get("canonical_family_slug", ""), row.get("canonical_subservice_slug", ""), row.get("exact_task_profile", ""),
        row.get("risk_level", ""), row.get("mandatory_bundles", ""), row.get("conditional_bundles", ""), row.get("subject_bindings", ""),
        row.get("policy_decision", ""), row.get("source_state", ""), row.get("legal_state", ""), "", "", "", "", "", "", "",
    ] for row in matrix_rows]
    write_sheet(review_book, "62 Coverage Review", coverage_review_headers, coverage_review_rows)
    source_review_headers = [
        "source_key", "authority_issuer", "law_or_regulation_title_as_captured", "section_article_class_exception", "effective_date_text", "official_url", "archive_reference", "archive_file_sha256", "current_source_status", "bound_requirement_bundles", "counsel_verified_title_blank", "counsel_verified_section_blank", "counsel_verified_effective_date_blank", "counsel_verified_exception_blank", "counsel_reviewer_identity_blank", "counsel_evidence_hash_blank", "immutable_source_approval_ref_blank",
    ]
    source_review_rows = [[
        row.get("source_key", ""), row.get("authority_issuer", ""), row.get("law_or_regulation_title_as_captured", ""), row.get("section_article_class_exception", ""),
        row.get("effective_date_text", ""), row.get("official_url", ""), row.get("archive_reference", ""), row.get("archive_file_sha256", ""),
        row.get("source_status", ""), row.get("bound_requirement_bundles", ""), "", "", "", "", "", "", "",
    ] for row in source_rows]
    write_sheet(review_book, "28 Source Review", source_review_headers, source_review_rows)
    review_book.save(EXPORTS / "US_CA_LA_LEGAL_REVIEW_TEMPLATE.xlsx")

    connector_lines = [
        "# US-CA-LOS_ANGELES Connector Gap Report",
        "",
        "**Status:** 28/28 connector routes are `NOT_CONFIGURED` or otherwise non-operational. This report does not claim an official API, contract, permission or authority verification route.",
        "",
        "| Connector | Target registry/API | Current status | Permission/evidence | Returned fields | Expiry/revocation | NO-GO reason |",
        "|---|---|---|---|---|---|---|",
    ]
    for connector in connectors:
        connector_lines.append(
            "| {key} | {target} | {status} / {assurance} | {permission} | {fields} | {expiry} | {reason} |".format(
                key=connector.get("connectorKey", ""), target=connector.get("targetRegistryOrApi", ""), status=connector.get("status", ""), assurance=connector.get("assuranceLevel", ""),
                permission=connector.get("accessPermissionRequirement", ""), fields=connector.get("returnedFields", ""), expiry=connector.get("expirySuspensionRevocationCheck", ""), reason=connector.get("noGoReason", "").replace("|", "/"),
            )
        )
    connector_lines.extend([
        "",
        "> Public webpages, OCR and AI extraction are not official verification connectors. `forbiddenScraping=1` remains in force for every route until an authorized integration is independently evidenced.",
    ])
    (EXPORTS / "US_CA_LA_CONNECTOR_GAP_REPORT.md").write_text("\n".join(connector_lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
